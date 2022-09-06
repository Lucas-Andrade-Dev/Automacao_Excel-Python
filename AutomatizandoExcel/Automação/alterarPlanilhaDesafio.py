from cmath import e
import openpyxl
from openpyxl.styles import PatternFill, Font

book = openpyxl.load_workbook('Eletronicos.xlsx')
eletronicos = book['Eletronicos']

my_fill = PatternFill(start_color='5399FF',end_color='5399FF', fill_type='solid')
my_font = Font(bold=True)
my_header = ['A1', 'B1', 'C1','D1','E1'] 
for cell in my_header: 
    eletronicos[cell].fill = my_fill 
    eletronicos[cell].font = my_font
eletronicos['E1'] = 'Total'
for i in range(2,5):
    eletronicos['E' + str(i)] = f'=(C{i}*D{i})' 
    eletronicos['E' + str(i)].font = my_font 
    eletronicos['E' + str(i)].fill = my_fill
eletronicos['E5'] = f'=SUM(E2:E4)'

book.save('Eletronicos.xlsx')



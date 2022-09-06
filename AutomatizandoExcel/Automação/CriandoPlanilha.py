import openpyxl

#criar uma planilha(book)
book = openpyxl.Workbook()
# como visualizar paginas existentes
print(book.sheetnames)
#como criar uma pagina
book.create_sheet('Orçamento')
#como selecionar uma pagina
orcamento_page = book['Orçamento']
#criando colunar
orcamento_page.append(['Celular','QTD','Preço'])
orcamento_page.append(['Xiaomi mi8','2','R$1499,99'])
orcamento_page.append(['Samsung A10','5','R$999,99'])
orcamento_page.append(['Xiaomi note 10','1','R$2499,99'])
orcamento_page.append(['Moto e5','6','R$1350,00'])
#para salvar
book.save('PlanilhasDeOrçamentos.xlsx')

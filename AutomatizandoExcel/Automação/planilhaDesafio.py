import openpyxl
from msilib.schema import Font

from openpyxl.styles import Font 
from openpyxl.styles import Border
from openpyxl.styles import PatternFill, Font


book = openpyxl.Workbook()
teste = book.active

book.create_sheet('Eletronicos')
eletronicos = book['Eletronicos']
my_fill = PatternFill(start_color='5399FF',  end_color='5399FF', fill_type='solid')
my_font = Font(bold=True)
eletronicos.column_dimensions['A'].width = 20
eletronicos.column_dimensions['B'].width = 20
eletronicos.column_dimensions['C'].width = 20
eletronicos.column_dimensions['D'].width = 20
eletronicos.column_dimensions['E'].width = 20

eletronicos.append(['Eletronica','Memoria','QTD','Preço'])
eletronicos.append(['Xiaomi mi 8','8gb ram',2, 2500])
eletronicos.append(['Samsung A10','16gb ram',3,5500])
eletronicos.append(['Moto e9','32gb ram',4,8500])
book.save('Eletronicos.xlsx')